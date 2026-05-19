"""Import customer profile Excel into MySQL dimension tables.

This script keeps two tables:
- dim_customer_profile: current customer profile used for fast joins.
- dim_customer_profile_snapshot: one row per import batch and customer.
"""

import argparse
import datetime as dt
import hashlib
import json
import re
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import pymysql
from openpyxl import load_workbook


COLUMN_DEFS: List[Tuple[str, str]] = [
    ("cust_id", "VARCHAR(64) NOT NULL"),
    ("license_code", "VARCHAR(64) NULL"),
    ("shop_id", "VARCHAR(64) NULL"),
    ("tenant_id", "VARCHAR(64) NULL"),
    ("cust_name", "VARCHAR(255) NULL"),
    ("shop_name", "VARCHAR(255) NULL"),
    ("linkman", "VARCHAR(128) NULL"),
    ("order_phone", "VARCHAR(64) NULL"),
    ("cust_seg_name", "VARCHAR(64) NULL"),
    ("address", "VARCHAR(500) NULL"),
    ("cigar_grade", "VARCHAR(64) NULL"),
    ("sale_dept", "VARCHAR(128) NULL"),
    ("ss_name", "VARCHAR(128) NULL"),
    ("slsman", "VARCHAR(128) NULL"),
    ("net_date", "DATE NULL"),
    ("base_type_name", "VARCHAR(128) NULL"),
    ("base_type_sub_name", "VARCHAR(128) NULL"),
    ("market_type", "VARCHAR(64) NULL"),
    ("work_port_name", "VARCHAR(64) NULL"),
    ("business_area_type", "VARCHAR(128) NULL"),
    ("terminal_type", "VARCHAR(128) NULL"),
    ("modern_terminal_sub_type", "VARCHAR(128) NULL"),
    ("terminal_sub_type", "VARCHAR(128) NULL"),
    ("terminal_level", "VARCHAR(64) NULL"),
    ("scan_device", "VARCHAR(128) NULL"),
    ("business_status", "VARCHAR(64) NULL"),
    ("order_method", "VARCHAR(128) NULL"),
    ("order_cycle_type", "VARCHAR(64) NULL"),
    ("order_week", "VARCHAR(64) NULL"),
    ("order_day", "VARCHAR(64) NULL"),
    ("bank_name", "VARCHAR(128) NULL"),
    ("bank_account", "VARCHAR(128) NULL"),
    ("bank_account_name", "VARCHAR(128) NULL"),
    ("short_name", "VARCHAR(255) NULL"),
    ("mnemonic_code", "VARCHAR(128) NULL"),
    ("is_railway", "VARCHAR(16) NULL"),
    ("is_chain", "VARCHAR(16) NULL"),
    ("is_order_paused", "VARCHAR(16) NULL"),
    ("special_tag", "VARCHAR(255) NULL"),
    ("business_area_sub_type", "VARCHAR(128) NULL"),
    ("structure_category", "VARCHAR(128) NULL"),
    ("payment_method", "VARCHAR(128) NULL"),
    ("is_online_payment", "VARCHAR(16) NULL"),
    ("cloud_pos_status", "VARCHAR(128) NULL"),
    ("longitude", "DECIMAL(12,8) NULL"),
    ("latitude", "DECIMAL(12,8) NULL"),
    ("invoice_type", "VARCHAR(128) NULL"),
    ("group_name", "VARCHAR(128) NULL"),
    ("standard_terminal_type", "VARCHAR(128) NULL"),
    ("business_area_ext", "VARCHAR(128) NULL"),
    ("special_group", "VARCHAR(128) NULL"),
    ("cigar_terminal_type", "VARCHAR(128) NULL"),
    ("is_three_full_store", "VARCHAR(16) NULL"),
    ("invoice_name", "VARCHAR(255) NULL"),
    ("is_price_sample", "VARCHAR(16) NULL"),
    ("business_scale", "VARCHAR(64) NULL"),
    ("special_market", "VARCHAR(128) NULL"),
    ("special_market_sub", "VARCHAR(128) NULL"),
    ("row_hash", "CHAR(64) NULL"),
    ("source_file", "VARCHAR(500) NULL"),
    ("snapshot_date", "DATE NULL"),
    ("import_batch_id", "VARCHAR(32) NOT NULL"),
    ("updated_at", "DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP"),
    ("etl_loaded_at", "DATETIME NOT NULL DEFAULT CURRENT_TIMESTAMP"),
]


EXCEL_TO_DB = {
    "许可证号": "license_code",
    "客户编码": "cust_id",
    "客户名称": "cust_name",
    "负责人": "linkman",
    "订货电话": "order_phone",
    "客户档位": "cust_seg_name",
    "经营地址": "address",
    "雪茄挡位": "cigar_grade",
    "区县": "sale_dept",
    "市场部": "ss_name",
    "营销线路": "slsman",
    "入网日期": "net_date",
    "经营业态": "base_type_name",
    "经营业态细分": "base_type_sub_name",
    "市场类型": "market_type",
    "城乡分类": "work_port_name",
    "商圈类型": "business_area_type",
    "终端类型": "terminal_type",
    "现代终端细分": "modern_terminal_sub_type",
    "终端细分": "terminal_sub_type",
    "终端等级": "terminal_level",
    "扫码设备": "scan_device",
    "经营状态": "business_status",
    "订货方式": "order_method",
    "订货周期类型": "order_cycle_type",
    "订货周次": "order_week",
    "订货日": "order_day",
    "开户行": "bank_name",
    "银行帐号": "bank_account",
    "开户姓名": "bank_account_name",
    "简称": "short_name",
    "助记码": "mnemonic_code",
    "是否铁路户": "is_railway",
    "是否连锁": "is_chain",
    "是否暂停订货": "is_order_paused",
    "特殊标签": "special_tag",
    "商圈类型细分": "business_area_sub_type",
    "结构类别": "structure_category",
    "支付方式": "payment_method",
    "是否网上支付": "is_online_payment",
    "云pos状态": "cloud_pos_status",
    "经度": "longitude",
    "纬度": "latitude",
    "发票类型": "invoice_type",
    "所属小组": "group_name",
    "标准终端类型": "standard_terminal_type",
    "商圈类型(扩展)": "business_area_ext",
    "特殊群体": "special_group",
    "雪茄终端类型": "cigar_terminal_type",
    "是否三全门店": "is_three_full_store",
    "发票名称": "invoice_name",
    "是否价采户": "is_price_sample",
    "经营规模": "business_scale",
    "特类市场": "special_market",
    "特类市场细分": "special_market_sub",
}


LEGACY_CURRENT_COLUMN_MAP = {
    "custId": "cust_id",
    "licenseNo": "license_code",
    "custName": "cust_name",
    "principalName": "linkman",
    "orderPhone": "order_phone",
    "customerSegment": "cust_seg_name",
    "businessAddress": "address",
    "cigarSegment": "cigar_grade",
    "countyDistrict": "sale_dept",
    "marketDepartment": "ss_name",
    "marketingRoute": "slsman",
    "networkJoinDate": "net_date",
    "businessType": "base_type_name",
    "businessTypeDetail": "base_type_sub_name",
    "marketType": "market_type",
    "urbanRuralCategory": "work_port_name",
    "businessCircleType": "business_area_type",
    "terminalType": "terminal_type",
    "modernTerminalDetail": "modern_terminal_sub_type",
    "terminalDetail": "terminal_sub_type",
    "terminalLevel": "terminal_level",
    "scanDevice": "scan_device",
    "businessStatus": "business_status",
    "orderMethod": "order_method",
    "orderCycleType": "order_cycle_type",
    "orderWeek": "order_week",
    "orderDay": "order_day",
    "bankName": "bank_name",
    "bankAccount": "bank_account",
    "accountHolderName": "bank_account_name",
    "shortName": "short_name",
    "mnemonicCode": "mnemonic_code",
    "isRailwayCustomer": "is_railway",
    "isChainStore": "is_chain",
    "isOrderSuspended": "is_order_paused",
    "specialTag": "special_tag",
    "businessCircleTypeDetail": "business_area_sub_type",
    "structureCategory": "structure_category",
    "paymentMethod": "payment_method",
    "isOnlinePayment": "is_online_payment",
    "cloudPosStatus": "cloud_pos_status",
    "invoiceType": "invoice_type",
    "belongingGroup": "group_name",
    "standardTerminalType": "standard_terminal_type",
    "businessCircleTypeExt": "business_area_ext",
    "specialPopulation": "special_group",
    "cigarTerminalType": "cigar_terminal_type",
    "isSanQuanStore": "is_three_full_store",
    "invoiceName": "invoice_name",
    "isPriceProcurementCustomer": "is_price_sample",
    "businessScale": "business_scale",
    "specialMarketType": "special_market",
    "specialMarketTypeDetail": "special_market_sub",
    "sourceFileName": "source_file",
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_cust_id(value: Any) -> str:
    text = normalize_text(value)
    return re.sub(r"\{.*?\}", "", text).strip()


def to_date(value: Any) -> Optional[dt.date]:
    if value in (None, ""):
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = normalize_text(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y%m%d"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def to_decimal_text(value: Any) -> Optional[str]:
    if value in (None, ""):
        return None
    text = normalize_text(value)
    return text or None


def find_excel_path(raw_path: str) -> Path:
    path = Path(raw_path)
    if path.exists():
        return path
    root = Path.cwd()
    candidates = list(root.glob("*.xlsx"))
    if not candidates:
        candidates = list(root.parent.glob("*.xlsx"))
    for candidate in candidates:
        if "客户信息查询" in candidate.name:
            return candidate
    raise FileNotFoundError(raw_path)


def get_columns() -> List[str]:
    return [name for name, _ in COLUMN_DEFS]


def column_def_map() -> Dict[str, str]:
    return dict(COLUMN_DEFS)


def ensure_database(conn: Any, database: str) -> None:
    with conn.cursor() as cursor:
        cursor.execute(
            f"CREATE DATABASE IF NOT EXISTS `{database}` "
            "DEFAULT CHARACTER SET utf8mb4 COLLATE utf8mb4_general_ci"
        )
        cursor.execute(f"USE `{database}`")


def existing_columns(conn: Any, table_name: str) -> set:
    with conn.cursor() as cursor:
        cursor.execute("SHOW COLUMNS FROM `{}`".format(table_name))
        return {row[0] for row in cursor.fetchall()}


def existing_indexes(conn: Any, table_name: str) -> set:
    with conn.cursor() as cursor:
        cursor.execute("SHOW INDEX FROM `{}`".format(table_name))
        return {row[2] for row in cursor.fetchall()}


def ensure_current_table(conn: Any) -> None:
    defs = column_def_map()
    with conn.cursor() as cursor:
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS dim_customer_profile (
              cust_id VARCHAR(64) NOT NULL,
              PRIMARY KEY (cust_id)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """
        )
    current_cols = existing_columns(conn, "dim_customer_profile")
    with conn.cursor() as cursor:
        for name, definition in defs.items():
            if name not in current_cols:
                cursor.execute(f"ALTER TABLE dim_customer_profile ADD COLUMN `{name}` {definition}")
    add_indexes(conn, "dim_customer_profile")


def ensure_snapshot_table(conn: Any) -> None:
    defs = ",\n              ".join(f"`{name}` {definition}" for name, definition in COLUMN_DEFS)
    with conn.cursor() as cursor:
        cursor.execute(
            f"""
            CREATE TABLE IF NOT EXISTS dim_customer_profile_snapshot (
              {defs},
              PRIMARY KEY (import_batch_id, cust_id),
              KEY idx_snapshot_cust_id (cust_id),
              KEY idx_snapshot_license_code (license_code),
              KEY idx_snapshot_date (snapshot_date),
              KEY idx_snapshot_seg_date (cust_seg_name, snapshot_date)
            ) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4
            """
        )


def add_indexes(conn: Any, table_name: str) -> None:
    indexes = existing_indexes(conn, table_name)
    index_sql = {
        "idx_dim_license_code": f"CREATE INDEX idx_dim_license_code ON `{table_name}` (license_code)",
        "idx_dim_shop_id": f"CREATE INDEX idx_dim_shop_id ON `{table_name}` (shop_id)",
        "idx_dim_sale_org": f"CREATE INDEX idx_dim_sale_org ON `{table_name}` (sale_dept, ss_name, slsman)",
        "idx_dim_cust_seg": f"CREATE INDEX idx_dim_cust_seg ON `{table_name}` (cust_seg_name)",
        "idx_dim_base_type": f"CREATE INDEX idx_dim_base_type ON `{table_name}` (base_type_name)",
        "idx_dim_updated_at": f"CREATE INDEX idx_dim_updated_at ON `{table_name}` (updated_at)",
    }
    with conn.cursor() as cursor:
        for index_name, sql in index_sql.items():
            if index_name not in indexes:
                cursor.execute(sql)


def row_hash(row: Dict[str, Any]) -> str:
    stable = {key: row.get(key) for key in sorted(row) if key not in {"updated_at", "etl_loaded_at"}}
    payload = json.dumps(stable, ensure_ascii=False, sort_keys=True, default=str)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def read_rows(path: Path, snapshot_date: dt.date, import_batch_id: str) -> List[Dict[str, Any]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    raw_headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [normalize_text(value) for value in raw_headers]
    db_columns = get_columns()
    rows: List[Dict[str, Any]] = []

    for values in sheet.iter_rows(min_row=2, values_only=True):
        source = dict(zip(headers, values))
        row = {column: None for column in db_columns}
        for excel_name, db_name in EXCEL_TO_DB.items():
            value = source.get(excel_name)
            if db_name == "cust_id":
                row[db_name] = clean_cust_id(value)
            elif db_name == "net_date":
                row[db_name] = to_date(value)
            elif db_name in {"longitude", "latitude"}:
                row[db_name] = to_decimal_text(value)
            else:
                row[db_name] = normalize_text(value) or None

        if not row["cust_id"]:
            continue
        row["shop_id"] = row.get("license_code")
        row["tenant_id"] = row.get("license_code")
        row["shop_name"] = row.get("cust_name")
        row["source_file"] = str(path)
        row["snapshot_date"] = snapshot_date
        row["import_batch_id"] = import_batch_id
        row["row_hash"] = row_hash(row)
        rows.append(row)

    return rows


def chunks(items: List[Dict[str, Any]], size: int) -> Iterable[List[Dict[str, Any]]]:
    for start in range(0, len(items), size):
        yield items[start : start + size]


def value_for_column(row: Dict[str, Any], column: str) -> Any:
    if column in row:
        return row.get(column)
    mapped = LEGACY_CURRENT_COLUMN_MAP.get(column)
    if mapped:
        return row.get(mapped)
    return None


def upsert_current(conn: Any, rows: List[Dict[str, Any]]) -> None:
    current_cols = existing_columns(conn, "dim_customer_profile")
    base_cols = [col for col in get_columns() if col not in {"updated_at", "etl_loaded_at"} and col in current_cols]
    legacy_cols = [col for col in LEGACY_CURRENT_COLUMN_MAP if col in current_cols and col not in base_cols]
    insert_cols = legacy_cols + base_cols
    col_sql = ", ".join(f"`{col}`" for col in insert_cols)
    placeholders = ", ".join(["%s"] * len(insert_cols))
    updates = ", ".join(
        f"`{col}`=VALUES(`{col}`)"
        for col in insert_cols
        if col not in {"custId", "cust_id"}
    )
    sql = (
        f"INSERT INTO dim_customer_profile ({col_sql}) VALUES ({placeholders}) "
        f"ON DUPLICATE KEY UPDATE {updates}, updated_at=NOW()"
    )
    payload = [tuple(value_for_column(row, col) for col in insert_cols) for row in rows]
    with conn.cursor() as cursor:
        cursor.executemany(sql, payload)


def insert_snapshot(conn: Any, rows: List[Dict[str, Any]]) -> None:
    columns = get_columns()
    insert_cols = [col for col in columns if col not in {"updated_at", "etl_loaded_at"}]
    col_sql = ", ".join(f"`{col}`" for col in insert_cols)
    placeholders = ", ".join(["%s"] * len(insert_cols))
    updates = ", ".join(
        f"`{col}`=VALUES(`{col}`)"
        for col in insert_cols
        if col not in {"import_batch_id", "cust_id"}
    )
    sql = (
        f"INSERT INTO dim_customer_profile_snapshot ({col_sql}) VALUES ({placeholders}) "
        f"ON DUPLICATE KEY UPDATE {updates}, updated_at=NOW()"
    )
    payload = [tuple(row.get(col) for col in insert_cols) for row in rows]
    with conn.cursor() as cursor:
        cursor.executemany(sql, payload)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Import customer profile Excel into MySQL.")
    parser.add_argument("--excel", required=True, help="Path to customer profile Excel")
    parser.add_argument("--host", default="127.0.0.1")
    parser.add_argument("--port", type=int, default=3306)
    parser.add_argument("--user", default="root")
    parser.add_argument("--password", default="")
    parser.add_argument("--database", default="pos_ods")
    parser.add_argument("--snapshot-date", default=dt.date.today().strftime("%Y-%m-%d"))
    parser.add_argument("--batch-size", type=int, default=1000)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    excel_path = find_excel_path(args.excel)
    snapshot_date = dt.datetime.strptime(args.snapshot_date, "%Y-%m-%d").date()
    import_batch_id = dt.datetime.now().strftime("%Y%m%d%H%M%S")

    conn = pymysql.connect(
        host=args.host,
        port=args.port,
        user=args.user,
        password=args.password,
        charset="utf8mb4",
        autocommit=False,
    )
    try:
        ensure_database(conn, args.database)
        ensure_current_table(conn)
        ensure_snapshot_table(conn)
        conn.commit()

        rows = read_rows(excel_path, snapshot_date, import_batch_id)
        for part in chunks(rows, args.batch_size):
            upsert_current(conn, part)
            insert_snapshot(conn, part)
            conn.commit()

        print(f"import_batch_id={import_batch_id}")
        print(f"source_file={excel_path}")
        print(f"snapshot_date={snapshot_date}")
        print(f"imported_rows={len(rows)}")
        print("target_tables=dim_customer_profile, dim_customer_profile_snapshot")
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
